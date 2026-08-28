from pathlib import Path
from openpyxl import load_workbook

source = Path('/home/ubuntu/upload/farmaville-catalogo-produtos.xlsx')
wb = load_workbook(source, read_only=True, data_only=True)
print('SHEETS', wb.sheetnames)
for ws in wb.worksheets:
    print(f'\nSHEET {ws.title} rows={ws.max_row} cols={ws.max_column}')
    rows = ws.iter_rows(values_only=True)
    for index, row in zip(range(6), rows):
        print(index + 1, list(row))
